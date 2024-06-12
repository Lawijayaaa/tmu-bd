import threading
import time
import openpyxl
import logging
import os, sys
from toolboxTMU import initTkinter
from openpyxl import Workbook
from pymodbus.client import ModbusSerialClient

#init value
engineName = " Trafo X"
progStat = True

#init logger sysdata
ts = time.strftime("%Y%m%d")
pathStrWin = r'C:\Users\Lutfi.LUTFI-PC\Desktop\tmu-bd\assets\sysdata\syslog-'
pathStrUnixTest = r'/home/pi/tmu-bd/assets/sysdata Test/syslog-'
#pathStrUnixFinal = r'/home/pi/tmu-bd/assets/sysdata Test/syslog-'
pathSysLog = pathStrUnixTest + ts + engineName + '.log'
logging.basicConfig(filename=pathSysLog, format='[%(asctime)s] | %(levelname)s: %(message)s',level=logging.INFO)

#init logger rawdata
pathStrWin = r'C:\Users\Lutfi.LUTFI-PC\Desktop\tmu-bd\assets\rawdata\datalogger-'
pathStrUnixTest = r'/home/pi/tmu-bd/assets/rawdata Test/datalogger-'
#pathStrUnixFinal = r'/home/pi/tmu-bd/assets/rawdata/datalogger-'
pathDatLog = pathStrUnixTest + ts + engineName + '.xlsx'

#init modbus device
client = ModbusSerialClient(method = 'rtu', port = '/dev/ttyACM0', baudrate = 19200)

try:
    wb = openpyxl.load_workbook(pathDatLog)
    logging.info("Open Existing Datalog")
except:
    #create new datalog
    logging.info("Create New Datalog")
    workbook = Workbook()
    workbook.save(pathDatLog)
    #create datalog's header
    wb = openpyxl.load_workbook(pathDatLog)
    sheet = wb.active
    sheet.title = "Raw_data"
    name = (('timestamp', 'OilTemp', 'BusTemp1', 'BusTemp2', 'BusTemp3', 'Press', 'Level',
            'Van', 'Vab', 'Ia', 'PFa', 'Pa', 'Qa', 'Sa', 'THDV1', 'THDI1',
            'Vbn', 'Vbc', 'Ib', 'PFb', 'Pb', 'Qb', 'Sb', 'THDV2', 'THDI2',
            'Vcn', 'Vca', 'Ic', 'PFc', 'Pc', 'Qc', 'Sc', 'THDV3', 'THDI3',
            'Itot', 'PFsyst', 'Psig', 'Qsig', 'Ssig', 'Freq', 'Ineutral',
            'kWhInp', 'kWhOut', 'kVARhinp', 'kVARhOut',
            'KRateda', 'deRatinga', 'KRatedb', 'deRatingb', 'KRatedc', 'deRatingc',
            'WTITemp1', 'WTITemp2', 'WTITemp3', 
            'trafoStatus', 'DI stat'),)
    for row in name:
        sheet.append(row)
    sheetName = ["Harmonic_phR", "Harmonic_phS", "Harmonic_phT"]
    for member in sheetName:
        wb.create_sheet(member)
    for name in sheetName:
        sheetHarm = wb[name]
        rows = (('timestamp', 'V 1st', 'V 3rd' , 'V 5th' , 'V 7th' , 'V 9th' , 'V 11th' , 'V 13th' , 'V 15th' ,
                 'V 17th' , 'V 19th' , 'V 21st' , 'V 23rd' , 'V 25th' , 'V 27th' , 'V 29th' , 'V 31st',
                 'I 1st', 'I 3rd' , 'I 5th' , 'I 7th' , 'I 9th' , 'I 11th' , 'I 13th' , 'I 15th' ,
                 'I 17th' , 'I 19th' , 'I 21st' , 'I 23rd' , 'I 25th' , 'I 27th' , 'I 29th' , 'I 31st'),)
        for row in rows:
            sheetHarm.append(row)
    wb.save(pathDatLog)

#main program
def mainloop(thread_name, interval):
    logging.info("Program Started")
    global progStat
    #main loop
    while True:
        while progStat:
            mainScreen.progStatLbl['text'] = "Running"
            time.sleep(2)
        else:
            mainScreen.progStatLbl["text"] = "Stop"
            time.sleep(2)

#restart program
def Restart():
    global wb
    logging.info("Saving Excel File before Restart")
    wb.save(pathDatLog)
    os.execv(sys.executable, [sys.executable] + ['/home/pi/tmu-bd/main.py'])

#continue program
def Start():
    global progStat
    progStat = True

#pause program
def Stop():
    global progStat
    progStat = False

if __name__ == "__main__":
    #import tkinter, connect buttons with function
    mainScreen = initTkinter()
    mainScreen.restartBtn["command"] = Restart
    mainScreen.startBtn["command"] = Start
    mainScreen.stopBtn["command"] = Stop
    #threading to multi process between tkinter and main program
    thread1 = threading.Thread(target=mainloop, args=('thread1', 1))
    thread1.start()
    #start tkinter
    mainScreen.screen.mainloop()
